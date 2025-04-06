#TO WORK WITH THIS CODE WE NEED TO INSTALL (openpyxl library....first using pip install openpyxl command)

import openpyxl
wb= openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "Hello"
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "World"
c3 = sheet['A2']
c3.value = "Welcome"
c4 = sheet['B2']
c4.value = "Everyone"
wb.save("sample.xlsx")

#After this code get executed a 'sample.xlsx' excel file will be created